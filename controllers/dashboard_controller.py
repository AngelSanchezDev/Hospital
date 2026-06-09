import os
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from database.conexion import conexion


class DashboardController:
    def __init__(self):
        self._conexion = conexion
        self._validar_conexion()

    def _validar_conexion(self):
        try:
            if not self._conexion:
                raise Exception("Conexión no disponible")
            cursor = self._conexion.cursor()
            cursor.execute("SELECT 1")
        except Exception as e:
            raise Exception(f"Error validando conexión: {str(e)}")

    def _cursor(self):
        return self._conexion.cursor()

    def _fetch_one(self, sql, params=None):
        cursor = self._cursor()
        cursor.execute(sql, params or ())
        return cursor.fetchone()

    def _fetch_all(self, sql, params=None):
        cursor = self._cursor()
        cursor.execute(sql, params or ())
        return cursor.fetchall()

    def _zero_if_none(self, value):
        return 0.0 if value is None else float(value)

    def _scalar(self, sql, params=None):
        row = self._fetch_one(sql, params)
        return row[0] if row else 0

    @staticmethod
    def _format_date_offset(days):
        return (date.today() + timedelta(days=days)).strftime("%d/%m")

    # ── Indicadores KPI ─────────────────────────────────────────────────
    def obtener_indicadores(self):
        return {
            "facturado_hoy": self._zero_if_none(self._scalar(
                "SELECT ISNULL(SUM(total),0) FROM Factura "
                "WHERE eliminado=0 AND CAST(fecha_factura AS DATE)=CAST(GETDATE() AS DATE)"
            )),
            "facturado_mes": self._zero_if_none(self._scalar(
                "SELECT ISNULL(SUM(total),0) FROM Factura "
                "WHERE eliminado=0 AND YEAR(fecha_factura)=YEAR(GETDATE()) "
                "AND MONTH(fecha_factura)=MONTH(GETDATE())"
            )),
            "cobrado_hoy": self._zero_if_none(self._scalar(
                "SELECT ISNULL(SUM(total),0) FROM Factura "
                "WHERE eliminado=0 AND estado_pago='PAGADO' "
                "AND CAST(fecha_factura AS DATE)=CAST(GETDATE() AS DATE)"
            )),
            "cobrado_mes": self._zero_if_none(self._scalar(
                "SELECT ISNULL(SUM(total),0) FROM Factura "
                "WHERE eliminado=0 AND estado_pago='PAGADO' "
                "AND YEAR(fecha_factura)=YEAR(GETDATE()) "
                "AND MONTH(fecha_factura)=MONTH(GETDATE())"
            )),
            "facturas_pendientes": self._zero_if_none(self._scalar(
                "SELECT COUNT(*) FROM Factura WHERE eliminado=0 AND estado_pago='PENDIENTE'"
            )),
            "reservaciones_activas": self._zero_if_none(self._scalar(
                "SELECT COUNT(*) FROM Reservacion "
                "WHERE eliminado=0 AND estado=1 AND fecha_cita>=CAST(GETDATE() AS DATE)"
            )),
            "citas_hoy": self._zero_if_none(self._scalar(
                "SELECT COUNT(*) FROM Reservacion "
                "WHERE eliminado=0 AND fecha_cita=CAST(GETDATE() AS DATE)"
            )),
            "total_pacientes": self._zero_if_none(self._scalar(
                "SELECT COUNT(*) FROM Paciente WHERE eliminado=0"
            )),
        }

    # ── Gráficos ────────────────────────────────────────────────────────
    def ingresos_por_dia(self):
        rows = self._fetch_all(
            "SELECT CAST(fecha_factura AS DATE) fecha, ISNULL(SUM(total),0) total "
            "FROM Factura WHERE eliminado=0 "
            "AND fecha_factura>=DATEADD(day,-6,CAST(GETDATE() AS DATE)) "
            "GROUP BY CAST(fecha_factura AS DATE) ORDER BY fecha"
        )
        labels = [self._format_date_offset(-i) for i in range(6, -1, -1)]
        values = [0.0] * 7
        date_index = {label: idx for idx, label in enumerate(labels)}
        if rows:
            for fecha, total in rows:
                key = fecha.strftime("%d/%m")
                if key in date_index:
                    values[date_index[key]] = float(total)
        return labels, values

    def pagos_por_tipo(self):
        rows = self._fetch_all(
            "SELECT tp.nombre, ISNULL(SUM(r.precio),0) total "
            "FROM Reservacion r JOIN TipoPago tp ON r.id_pago=tp.id_pago "
            "WHERE r.eliminado=0 "
            "AND r.fecha_cita>=DATEADD(month,-1,CAST(GETDATE() AS DATE)) "
            "GROUP BY tp.nombre ORDER BY total DESC"
        )
        return [(n, float(t)) for n, t in rows] if rows else []

    # ── Nuevos: últimas citas y facturas ────────────────────────────────
    def ultimas_citas_hoy(self):
        """Devuelve las citas programadas para hoy."""
        rows = self._fetch_all(
            "SELECT TOP 8 "
            "  r.hora_cita, "
            "  p.nombres + ' ' + p.apellidos AS paciente, "
            "  m.nombres + ' ' + m.apellidos AS medico, "
            "  e.nombre AS estado "
            "FROM Reservacion r "
            "JOIN Paciente p ON r.id_paciente = p.id_paciente "
            "JOIN Medico    m ON r.id_medico   = m.id_medico "
            "JOIN Estado    e ON r.id_estado   = e.id_estado "
            "WHERE r.eliminado = 0 "
            "AND r.fecha_cita = CAST(GETDATE() AS DATE) "
            "ORDER BY r.hora_cita ASC"
        )
        return rows or []

    def ultimas_facturas(self):
        """Devuelve las últimas 8 facturas emitidas."""
        rows = self._fetch_all(
            "SELECT TOP 8 "
            "  f.numero_factura, "
            "  p.nombres + ' ' + p.apellidos AS paciente, "
            "  f.total, "
            "  f.estado_pago, "
            "  CAST(f.fecha_factura AS DATE) AS fecha "
            "FROM Factura f "
            "JOIN Paciente p ON f.id_paciente = p.id_paciente "
            "WHERE f.eliminado = 0 "
            "ORDER BY f.id_factura DESC"
        )
        return rows or []

    # ── Export Excel ─────────────────────────────────────────────────────
    def exportar_excel(self):
        """Genera un archivo Excel con KPIs, citas del día y últimas facturas."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            wb = openpyxl.Workbook()

            # ── Estilos comunes ─────────────────────────────────────────
            azul  = PatternFill("solid", fgColor="1565C0")
            verde = PatternFill("solid", fgColor="22c55e")
            gris  = PatternFill("solid", fgColor="D0E8FF")
            thin  = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin")
            )

            def header_row(ws, row, headers, fill=azul):
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center")
                    c.border = thin

            def auto_width(ws):
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            # ═══════════════════════════════════════════════════════════
            # HOJA 1 — KPIs
            # ═══════════════════════════════════════════════════════════
            ws1 = wb.active
            ws1.title = "KPIs"
            ws1["A1"] = "Sistema Hospitalario — Indicadores"
            ws1["A1"].font = Font(bold=True, size=14, color="0F4C81")
            ws1["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws1["A2"].font = Font(italic=True, color="555555")

            indicadores = self.obtener_indicadores()
            nombres = {
                "facturado_hoy":       "Facturado Hoy (S/)",
                "facturado_mes":       "Facturado Mes (S/)",
                "cobrado_hoy":         "Cobrado Hoy (S/)",
                "cobrado_mes":         "Cobrado Mes (S/)",
                "facturas_pendientes": "Facturas Pendientes",
                "reservaciones_activas":"Reservaciones Activas",
                "citas_hoy":           "Citas Hoy",
                "total_pacientes":     "Total Pacientes",
            }
            header_row(ws1, 4, ["Indicador", "Valor"])
            for i, (key, label) in enumerate(nombres.items(), 5):
                val = indicadores.get(key, 0)
                ws1.cell(row=i, column=1, value=label).border = thin
                cell = ws1.cell(row=i, column=2, value=round(float(val), 2))
                cell.border = thin
                cell.alignment = Alignment(horizontal="right")
                if i % 2 == 0:
                    ws1.cell(row=i, column=1).fill = gris
                    cell.fill = gris
            auto_width(ws1)

            # ═══════════════════════════════════════════════════════════
            # HOJA 2 — Citas del día
            # ═══════════════════════════════════════════════════════════
            ws2 = wb.create_sheet("Citas Hoy")
            ws2["A1"] = f"Citas del día — {date.today().strftime('%d/%m/%Y')}"
            ws2["A1"].font = Font(bold=True, size=13, color="0F4C81")
            header_row(ws2, 3, ["Hora", "Paciente", "Médico", "Estado"])
            citas = self.ultimas_citas_hoy()
            for i, row in enumerate(citas, 4):
                for col, val in enumerate(row, 1):
                    c = ws2.cell(row=i, column=col, value=str(val) if val else "")
                    c.border = thin
                    if i % 2 == 0: c.fill = gris
            auto_width(ws2)

            # ═══════════════════════════════════════════════════════════
            # HOJA 3 — Últimas Facturas
            # ═══════════════════════════════════════════════════════════
            ws3 = wb.create_sheet("Últimas Facturas")
            ws3["A1"] = "Últimas 8 Facturas Emitidas"
            ws3["A1"].font = Font(bold=True, size=13, color="0F4C81")
            header_row(ws3, 3, ["N° Factura", "Paciente", "Total (S/)", "Estado", "Fecha"])
            facturas = self.ultimas_facturas()
            for i, row in enumerate(facturas, 4):
                for col, val in enumerate(row, 1):
                    c = ws3.cell(row=i, column=col, value=str(val) if val else "")
                    c.border = thin
                    if i % 2 == 0: c.fill = gris
            auto_width(ws3)

            # ═══════════════════════════════════════════════════════════
            # HOJA 4 — Ingresos por día
            # ═══════════════════════════════════════════════════════════
            ws4 = wb.create_sheet("Ingresos 7 días")
            ws4["A1"] = "Ingresos por Día — Últimos 7 días"
            ws4["A1"].font = Font(bold=True, size=13, color="0F4C81")
            header_row(ws4, 3, ["Fecha", "Ingreso (S/)"])
            labels, values = self.ingresos_por_dia()
            for i, (lbl, val) in enumerate(zip(labels, values), 4):
                ws4.cell(row=i, column=1, value=lbl).border = thin
                c = ws4.cell(row=i, column=2, value=round(val, 2))
                c.border = thin
                c.alignment = Alignment(horizontal="right")
                if i % 2 == 0:
                    ws4.cell(row=i, column=1).fill = gris
                    c.fill = gris
            auto_width(ws4)

            # Guardar
            reports_dir = Path(__file__).resolve().parent.parent / "reports" / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            filename = f"Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = reports_dir / filename
            wb.save(str(path))
            return str(path)

        except ImportError:
            raise Exception("La librería openpyxl no está instalada.")
        except Exception as e:
            raise Exception(f"Error generando Excel: {str(e)}")
